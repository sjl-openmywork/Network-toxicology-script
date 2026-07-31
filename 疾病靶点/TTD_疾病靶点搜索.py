"""
TTD 靶点搜索脚本 v3.1 — 终端交互模式
================================================================================
功能：根据疾病关键词，从 TTD (Therapeutic Target Database) REST API 检索靶点蛋白信息
数据源：https://ttd.idrblab.cn/ (SPA 后端 REST API)
API Base: https://ttd.idrblab.cn/api/ttd

=== 核心 API ===
  POST /api/ttd/postSearchTargetResult     → 靶点搜索
  GET  /api/ttd/getTargetGeneralInfo/{id}  → 靶点详情（UniProt、染色体位置等）
  GET  /api/ttd/getSearchTargetByDiseaseList/{topic} → 疾病分类列表

=== 交互菜单 ===
  1. 疾病关键词搜索（基本模式）—— 秒级返回，含靶点ID、基因名、疾病、药物
  2. 疾病关键词搜索（详细模式）—— 含 UniProt ID、染色体位置、功能描述等
  3. 疾病分类精确搜索 —— 按 ICD 编码精准检索
  4. 浏览疾病分类列表 —— 分页查看 TTD 所有疾病分类
  0. 退出
"""

import os
import sys
import time
import logging
from datetime import datetime

import openpyxl
import requests

# ============ 日志配置 ============
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger(__name__)

# ============ API 常量 ============
API_BASE = "https://ttd.idrblab.cn/api/ttd"
SEARCH_ENDPOINT = f"{API_BASE}/postSearchTargetResult"
DETAIL_ENDPOINT = f"{API_BASE}/getTargetGeneralInfo"
DISEASE_LIST_ENDPOINT = f"{API_BASE}/getSearchTargetByDiseaseList"
REQUEST_TIMEOUT = 30
DETAIL_DELAY = 0.2
DISEASE_PAGE_SIZE = 30

HEADERS = {
    "Content-Type": "application/json",
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
    ),
    "Referer": "https://ttd.idrblab.cn/",
}


# ===================================================================
#  API 调用
# ===================================================================

def build_session() -> requests.Session:
    session = requests.Session()
    session.headers.update(HEADERS)
    return session


def search_targets(session: requests.Session, keyword: str,
                   engine: str = "fullText", topic: str = "no-topic") -> list[dict]:
    payload = {"engine": engine, "keyword": keyword, "topic": topic}
    try:
        resp = session.post(SEARCH_ENDPOINT, json=payload, timeout=REQUEST_TIMEOUT)
        resp.raise_for_status()
        return resp.json().get("results", [])
    except requests.exceptions.Timeout:
        logger.error("搜索请求超时")
        return []
    except requests.exceptions.ConnectionError:
        logger.error("连接 TTD API 失败，请检查网络")
        return []
    except requests.exceptions.HTTPError as e:
        logger.error(f"HTTP 错误: {e.response.status_code}")
        return []
    except Exception as e:
        logger.error(f"未知错误: {type(e).__name__}: {e}")
        return []


def get_target_detail(session: requests.Session, target_id: str) -> dict | None:
    try:
        resp = session.get(f"{DETAIL_ENDPOINT}/{target_id}", timeout=REQUEST_TIMEOUT)
        resp.raise_for_status()
        return resp.json()
    except Exception as e:
        logger.warning(f"获取 {target_id} 详情失败: {e}")
        return None


def get_disease_categories(session: requests.Session) -> list[str]:
    try:
        resp = session.get(f"{DISEASE_LIST_ENDPOINT}/no-topic", timeout=REQUEST_TIMEOUT)
        resp.raise_for_status()
        return resp.json().get("results", [])
    except Exception as e:
        logger.error(f"获取疾病列表失败: {e}")
        return []


# ===================================================================
#  数据处理与保存
# ===================================================================

DB_NAME = "TTD"


def _safe_filename(keyword: str) -> str:
    """文件名安全处理：移除特殊字符，保留字母数字与词间空格"""
    safe = "".join(c for c in keyword if c.isalnum() or c in " _-()（）")
    return " ".join(safe.split()) or "ttd_output"


def _output_dir() -> str:
    """输出子文件夹（与脚本同名，如 ttp），不存在时自动创建"""
    script_name = os.path.splitext(os.path.basename(__file__))[0].lower()
    out = os.path.join(os.path.dirname(os.path.abspath(__file__)), script_name)
    os.makedirs(out, exist_ok=True)
    return out


def save_basic_results(results: list[dict], keyword: str) -> str:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filepath = os.path.join(
        _output_dir(), f"{DB_NAME}_{_safe_filename(keyword)}_{timestamp}.xlsx"
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "TTD靶点搜索"

    ws.append(["序号", "靶点 ID", "靶点名称", "基因名", "靶点类型", "生化分类", "相关疾病", "代表药物"])
    for idx, item in enumerate(results, start=1):
        drugs = item.get("representative_drug", []) or []
        drug_str = "; ".join(d[1] if isinstance(d, list) and len(d) > 1 else str(d) for d in drugs)
        ws.append([
            idx, item.get("target_id", ""), item.get("target_name", ""),
            item.get("gene_name", ""), item.get("target_type", ""),
            item.get("biochemical_class", ""), item.get("diseases", "") or "", drug_str,
        ])

    for col, w in zip("ABCDEFGH", [6, 12, 40, 28, 14, 28, 55, 40]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"
    wb.save(filepath)
    logger.info(f"文件已保存: {filepath}")
    return filepath


def save_detailed_results(results: list[dict], details: list[dict | None], keyword: str) -> str:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filepath = os.path.join(
        _output_dir(), f"{DB_NAME}_{_safe_filename(keyword)}_详细_{timestamp}.xlsx"
    )

    wb = openpyxl.Workbook()

    # Sheet 1: 靶点概览
    ws1 = wb.active
    ws1.title = "靶点概览"
    ws1.append(["序号", "靶点 ID", "靶点名称", "基因名", "靶点类型", "生化分类",
                 "相关疾病", "代表药物", "UniProt (Organism)", "染色体位置", "HGNC", "Ensembl"])
    for idx, (item, detail) in enumerate(zip(results, details), start=1):
        drugs = item.get("representative_drug", []) or []
        drug_str = "; ".join(d[1] if isinstance(d, list) and len(d) > 1 else str(d) for d in drugs)
        gi = (detail or {}).get("general_info", {}) if detail else {}
        ws1.append([
            idx, item.get("target_id", ""), item.get("target_name", ""),
            item.get("gene_name", ""), item.get("target_type", ""),
            item.get("biochemical_class", ""), item.get("diseases", "") or "", drug_str,
            gi.get("uniprot_id_organism", ""), gi.get("chromosomal_location", ""),
            gi.get("hgnc", ""), gi.get("ensembl", ""),
        ])
    for col, w in zip("ABCDEFGHIJKL", [6, 12, 40, 28, 14, 28, 55, 40, 24, 16, 14, 22]):
        ws1.column_dimensions[col].width = w
    ws1.freeze_panes = "A2"

    # Sheet 2: 靶点功能描述
    ws2 = wb.create_sheet("靶点功能描述")
    ws2.append(["序号", "靶点 ID", "靶点名称", "基因名", "功能描述"])
    for idx, (item, detail) in enumerate(zip(results, details), start=1):
        gi = (detail or {}).get("general_info", {}) if detail else {}
        ws2.append([idx, item.get("target_id", ""), item.get("target_name", ""),
                     item.get("gene_name", ""), gi.get("function", "")])
    for col, w in zip("ABCDE", [6, 12, 40, 28, 120]):
        ws2.column_dimensions[col].width = w
    ws2.freeze_panes = "A2"

    # Sheet 3: 疾病关联详情
    ws3 = wb.create_sheet("疾病关联详情")
    ws3.append(["序号", "靶点 ID", "靶点名称", "基因名", "关联疾病", "ICD-11"])
    for idx, (item, detail) in enumerate(zip(results, details), start=1):
        gi = (detail or {}).get("general_info", {}) if detail else {}
        disease_list = gi.get("diseases", []) or []
        if disease_list:
            for d in disease_list:
                ws3.append([idx, item.get("target_id", ""), item.get("target_name", ""),
                             item.get("gene_name", ""), d.get("disease_entry_target", ""),
                             d.get("icd11_target", "")])
        else:
            ws3.append([idx, item.get("target_id", ""), item.get("target_name", ""),
                         item.get("gene_name", ""), "", ""])
    for col, w in zip("ABCDEF", [6, 12, 40, 28, 45, 22]):
        ws3.column_dimensions[col].width = w
    ws3.freeze_panes = "A2"

    wb.save(filepath)
    logger.info(f"文件已保存: {filepath}")
    return filepath


def preview_results(results: list[dict], count: int = 5):
    print(f"  前 {min(count, len(results))} 个靶点预览：")
    print(f"  {'靶点ID':<10} {'基因名':<28} {'靶点名称':<48} {'疾病':<60}")
    print(f"  {'-'*146}")
    for item in results[:count]:
        print(f"  {item.get('target_id',''):<10} {item.get('gene_name',''):<28} "
              f"{item.get('target_name','')[:45]:<48} {(item.get('diseases','') or '')[:58]}")


# ===================================================================
#  交互菜单
# ===================================================================

def clear_screen():
    os.system("cls" if os.name == "nt" else "clear")


def press_enter():
    input("\n按 Enter 返回主菜单...")


def show_banner():
    clear_screen()
    print("╔" + "═" * 56 + "╗")
    print("║" + "   TTD 靶点搜索工具 v3.1".center(54) + "║")
    print("║" + "   数据源: https://ttd.idrblab.cn/ (REST API)".center(48) + "║")
    print("║" + "   author: shenjianlin".center(48) + "║")
    print("║" + "   site: git@github.com:sjl-openmywork/Network-toxicology-script.git".center(36) + "║")
    print("╚" + "═" * 56 + "╝")


def menu_search_basic(session: requests.Session):
    """菜单 [1] 疾病关键词搜索（基本模式）"""
    show_banner()
    print("\n  ▸ 疾病关键词搜索 —— 基本模式")
    print("  " + "─" * 54)
    print("  直接搜索靶点，秒级返回（不含 UniProt 等详字段）\n")

    keyword = input("  请输入疾病名称（如 Alzheimer、Diabetes）: ").strip()
    if not keyword:
        print("\n  ⚠ 关键词不能为空")
        press_enter()
        return

    print()
    results = search_targets(session, keyword, engine="fullText")
    if not results:
        print(f"  ✗ 未找到「{keyword}」相关靶点，请检查拼写")
        press_enter()
        return

    logger.info(f"共找到 {len(results)} 个靶点")
    filepath = save_basic_results(results, keyword)
    print(f"\n  ✓ 共 {len(results)} 个靶点")
    print(f"  ✓ 文件: {filepath}")
    print()
    preview_results(results)
    press_enter()


def menu_search_detailed(session: requests.Session):
    """菜单 [2] 疾病关键词搜索（详细模式）"""
    show_banner()
    print("\n  ▸ 疾病关键词搜索 —— 详细模式")
    print("  " + "─" * 54)
    print("  搜索靶点并逐条获取详情（UniProt ID / 染色体位置 / 功能描述）")
    print("  每个靶点约 0.2s，靶点多时请耐心等待\n")

    keyword = input("  请输入疾病名称（如 Alzheimer、Diabetes）: ").strip()
    if not keyword:
        print("\n  ⚠ 关键词不能为空")
        press_enter()
        return

    print()
    results = search_targets(session, keyword, engine="fullText")
    if not results:
        print(f"  ✗ 未找到「{keyword}」相关靶点")
        press_enter()
        return

    logger.info(f"共找到 {len(results)} 个靶点")

    # 数量确认
    if len(results) > 50:
        print(f"  ⚠ 共 {len(results)} 个靶点，详细模式预计耗时约 {len(results)*0.2:.0f} 秒")
        confirm = input("  是否继续？(y/n，默认 y): ").strip().lower()
        if confirm == "n":
            print("  已取消")
            press_enter()
            return

    print()
    logger.info(f"正在获取 {len(results)} 个靶点的详细信息...")
    details: list[dict | None] = []
    success_count = 0

    for idx, item in enumerate(results, start=1):
        tid = item.get("target_id", "")
        if not tid:
            details.append(None)
            continue

        detail = get_target_detail(session, tid)
        details.append(detail)
        if detail:
            success_count += 1
            gi = detail.get("general_info", {})
            print(f"\r  进度: [{idx}/{len(results)}] {tid}  {gi.get('uniprot_id_organism', 'N/A'):16s}  "
                  f"{gi.get('chromosomal_location', ''):12s}", end="")
        if idx < len(results):
            time.sleep(DETAIL_DELAY)

    print()
    logger.info(f"详情获取完成: {success_count}/{len(results)} 成功")
    filepath = save_detailed_results(results, details, keyword)
    print(f"\n  ✓ 共 {len(results)} 个靶点，成功获取 {success_count} 个详情")
    print(f"  ✓ 文件: {filepath}")
    print(f"  ✓ Excel 含 3 个 Sheet：靶点概览 | 功能描述 | 疾病关联详情")
    print()
    preview_results(results)
    press_enter()


def menu_search_disease_class(session: requests.Session):
    """菜单 [3] 疾病分类精确搜索"""
    show_banner()
    print("\n  ▸ 疾病分类精确搜索")
    print("  " + "─" * 54)
    print("  使用 TTD 的 ICD-11 疾病编码进行精确搜索")
    print("  格式: ICD编码: 疾病名（如 5A11: Type 2 diabetes mellitus）\n")
    print("  提示: 选菜单 [4] 可浏览所有可用疾病分类编码\n")

    keyword = input("  请输入（格式: ICD编码: 疾病名）: ").strip()
    if not keyword or ":" not in keyword:
        print("\n  ⚠ 请输入正确格式（如 5A11: Type 2 diabetes mellitus）")
        press_enter()
        return

    parts = keyword.split(":", 1)
    topic = parts[0].strip()

    print()
    results = search_targets(session, keyword, engine="diseaseClass", topic=topic)
    if not results:
        print(f"  ✗ 未找到相关靶点")
        press_enter()
        return

    logger.info(f"共找到 {len(results)} 个靶点")
    filepath = save_basic_results(results, keyword)
    print(f"\n  ✓ 共 {len(results)} 个靶点")
    print(f"  ✓ 文件: {filepath}")
    print()
    preview_results(results)
    press_enter()


def menu_list_diseases(session: requests.Session):
    """菜单 [4] 浏览疾病分类列表（分页）"""
    show_banner()
    print("\n  ▸ 浏览疾病分类列表")
    print("  " + "─" * 54)
    print("  正在从 TTD 获取疾病分类...\n")

    categories = get_disease_categories(session)
    if not categories:
        print("  ✗ 获取失败，请检查网络")
        press_enter()
        return

    total = len(categories)
    total_pages = (total + DISEASE_PAGE_SIZE - 1) // DISEASE_PAGE_SIZE
    page = 0

    while True:
        show_banner()
        start = page * DISEASE_PAGE_SIZE
        end = min(start + DISEASE_PAGE_SIZE, total)
        subset = categories[start:end]

        print(f"\n  疾病分类列表（共 {total} 个）第 {page + 1}/{total_pages} 页")
        print("  " + "─" * 54)
        for i, cat in enumerate(subset, start=start + 1):
            print(f"  {i:4d}. {cat}")
        print("  " + "─" * 54)

        print(f"\n  [N] 下一页  [P] 上一页  [B] 返回主菜单")
        choice = input("  > ").strip().lower()

        if choice == "n" and page < total_pages - 1:
            page += 1
        elif choice == "p" and page > 0:
            page -= 1
        elif choice == "b":
            break
        else:
            if choice not in ("n", "p", "b"):
                print("  ⚠ 无效选项")


# ===================================================================
#  主入口
# ===================================================================

def main():
    session = build_session()

    while True:
        show_banner()
        print()
        print("  ┌────────────────────────────────────────────────────┐")
        print("  │  [1] 疾病关键词搜索（基本模式）    秒级  │")
        print("  │  [2] 疾病关键词搜索（详细模式）  含UniProt │")
        print("  │  [3] 疾病分类精确搜索          ICD编码检索　│")
        print("  │  [4] 浏览疾病分类列表          分页查看　　│")
        print("  │  [0] 退出                                      │")
        print("  └────────────────────────────────────────────────────┘")

        choice = input("\n  请选择 [0-4]: ").strip()

        if choice == "1":
            menu_search_basic(session)
        elif choice == "2":
            menu_search_detailed(session)
        elif choice == "3":
            menu_search_disease_class(session)
        elif choice == "4":
            menu_list_diseases(session)
        elif choice == "0":
            show_banner()
            print("\n  感谢使用 TTD 靶点搜索工具，再见！\n")
            break
        else:
            print(f"\n  ⚠ 无效选项「{choice}」，请重新选择")
            time.sleep(1)


if __name__ == "__main__":
    main()
