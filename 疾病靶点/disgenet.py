"""
DisGeNET 疾病靶点检索脚本 v1.4
================================================================================
功能：根据疾病名称/UMLS ID，从 DisGeNET REST API 检索关联基因及评分
数据源：https://www.disgenet.com/
API: http://api.disgenet.com/api/v1/

=== 前置条件 ===
  须先申请 DisGeNET 免费学术许可证：
    1. 访问 https://disgenet.com/academic-apply
    2. 使用机构邮箱（.edu / .edu.cn 等）注册
    3. 审核通过后，在 https://disgenet.com/Profile-area 获取 API Key
    4. 首次运行脚本时输入 API Key，自动保存到本地 .disgenet_key 文件

=== 免费学术计划限制 ===
  - 仅 CURATED 数据源（11 个精选数据库）
  - 最多 100 页 × 100 条/页 = 10,000 条结果
  - 不支持 ALL 数据源（非文本挖掘数据）

=== API 端点 ===
  GET /entity/disease?disease_free_text_search_string={name}  → 疾病名称搜索
  GET /entity/disease?page_number={0..99}                     → 全量疾病列表（不传搜索词）
  GET /gda/summary?disease={ids}&source={db}&score={min,max}  → 疾病→基因关联
  GET /entity/disease?disease={id}                            → 疾病映射查询

=== 疾病标识符支持 ===
  UMLS (C0028754), MONDO, OMIM, MeSH, DO, EFO, ORDO, ICD9CM, ICD10, NCI, HPO

=== 交互菜单 ===
  1. 按疾病名称搜索 → 输入名称 → 展示全部匹配 → (可选)导出表格 → 选择 → 获取基因
  2. 直接输入疾病 ID → UMLS/MONDO/OMIM 等
  3. 下载全部疾病列表 → 自动分页抓取 CURATED 全量疾病元数据
  4. 下载 Disease Ontology → 从官方 OBO 解析疾病术语 + UMLS 映射
  5. 下载 MONDO 本体 → 从官方 OBO 解析疾病术语 + UMLS 映射
  0. 退出

=== v1.1 更新 ===
  1. 疾病名称检索结果不再限制前 10 项，展示全部匹配疾病
  2. 新增「疾病检索结果」可选导出功能（仅 Excel .xlsx）

=== v1.2 更新 ===
  统一输出机制：结果保存到与脚本同名子文件夹（disgenet/），
  命名格式 DisGeNET_{检索词}_{YYYYMMDD_HHMMSS}.xlsx（仅 Excel）

=== v1.3 更新 ===
  新增 [3] 下载全部疾病列表：通过 entity/disease 端点（不传搜索词），
  自动翻页抓取全部 18,480 种疾病元数据（学术许可上限 10,000 条）

=== v1.4 更新 ===
  新增 [4] Disease Ontology 下载 + [5] MONDO 本体下载：
  从官方 OBO 文件流式解析，提取所有疾病术语及 UMLS CUI 交叉映射
"""

import os
import sys
import time
import logging
import json
from datetime import datetime

import openpyxl
from openpyxl.utils import get_column_letter
import requests

# ============ 日志配置 ============
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger(__name__)

# ============ 常量配置 ============
API_BASE = "http://api.disgenet.com/api/v1/"
REQUEST_TIMEOUT = 30
PAGE_DELAY = 1.5           # 分页间隔（API 要求）
KEY_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".disgenet_key")

# 外部疾病本体 OBO 文件下载地址（免费，无需 API Key）
OBO_SOURCES = {
    "DO": {
        "name": "Disease Ontology",
        "url": "https://raw.githubusercontent.com/DiseaseOntology/HumanDiseaseOntology/main/src/ontology/doid.obo",
        "id_prefix": "DOID:",
    },
    "MONDO": {
        "name": "MONDO (Mondo Disease Ontology)",
        "url": "https://purl.obolibrary.org/obo/mondo.obo",
        "id_prefix": "MONDO:",
    },
}

# 学术计划可用的数据库源
ACADEMIC_SOURCES = [
    "CURATED", "CLINGEN", "CLINVAR", "CLINPGX", "GENCC",
    "PSYGENET", "UNIPROT", "ORPHANET", "MGD_HUMAN", "RGD_HUMAN",
]

# 常见疾病预设（UMLS ID，来自 DisGeNET）
PRESET_DISEASES = {
    "1":  {"id": "C0002395", "name": "Alzheimer's Disease", "vocab": "UMLS"},
    "2":  {"id": "C0011849", "name": "Diabetes Mellitus", "vocab": "UMLS"},
    "3":  {"id": "C0011860", "name": "Type 2 Diabetes Mellitus", "vocab": "UMLS"},
    "4":  {"id": "C0028754", "name": "Obesity", "vocab": "UMLS"},
    "5":  {"id": "C0036341", "name": "Schizophrenia", "vocab": "UMLS"},
    "6":  {"id": "C0027051", "name": "Myocardial Infarction", "vocab": "UMLS"},
    "7":  {"id": "C0006142", "name": "Breast Carcinoma", "vocab": "UMLS"},
    "8":  {"id": "C0023893", "name": "Liver Cirrhosis, Experimental", "vocab": "UMLS"},
    "9":  {"id": "C0033860", "name": "Psoriasis", "vocab": "UMLS"},
    "10": {"id": "C0032460", "name": "Polycystic Ovary Syndrome", "vocab": "UMLS"},
    "11": {"id": "C0023890", "name": "Hepatitis, Chronic", "vocab": "UMLS"},
    "12": {"id": "C0004096", "name": "Asthma", "vocab": "UMLS"},
}


# ===================================================================
#  API Key 管理
# ===================================================================

def load_api_key() -> str | None:
    """加载保存的 API Key"""
    if os.path.exists(KEY_FILE):
        try:
            with open(KEY_FILE, "r") as f:
                return f.read().strip()
        except Exception:
            pass
    return None


def save_api_key(key: str):
    """保存 API Key 到本地文件"""
    with open(KEY_FILE, "w") as f:
        f.write(key.strip())
    logger.info(f"API Key 已保存到 {KEY_FILE}")


def get_api_key() -> str:
    """获取 API Key（文件/环境变量/输入）"""
    key = load_api_key() or os.environ.get("DISGENET_API_KEY", "")
    if not key:
        print("  未找到 API Key")
        print("  请先申请 DisGeNET 学术许可证: https://disgenet.com/academic-apply")
        print("  审核通过后在 Profile Area 获取 API Key")
        key = input("\n  请输入 API Key (直接回车跳过): ").strip()
        if key:
            save_api_key(key)
    return key


# ===================================================================
#  API 调用
# ===================================================================

def _request(session: requests.Session, url: str, api_key: str) -> dict | None:
    """带认证和重试的 API 请求"""
    for attempt in range(3):
        try:
            resp = session.get(
                url,
                headers={
                    "Authorization": api_key,
                    "Accept": "application/json",
                },
                timeout=REQUEST_TIMEOUT,
            )
            if resp.status_code == 429:
                wait_time = int(resp.headers.get("x-rate-limit-retry-after-seconds", 5))
                logger.warning(f"速率限制，等待 {wait_time} 秒...")
                if wait_time > 300:
                    logger.error(f"速率限制等待时间过长（{wait_time}s），请稍后再试")
                    return None
                time.sleep(wait_time)
                continue
            if resp.status_code == 403:
                data = resp.json()
                err = data.get("detail", str(resp.text))
                logger.error(f"权限不足: {err}")
                return None
            if resp.status_code != 200:
                logger.error(f"HTTP {resp.status_code}: {resp.text[:200]}")
                return None
            return resp.json()
        except requests.exceptions.Timeout:
            logger.warning(f"请求超时，重试 {attempt + 1}/3...")
            time.sleep(2)
        except Exception as e:
            logger.error(f"请求异常: {e}")
            return None
    return None


def check_account_profile(session: requests.Session, api_key: str) -> str:
    """
    检查当前账户类型，返回 "TRIAL" / "ACADEMIC" / "UNKNOWN"

    TRIAL: 7天试用（限制：前30条，不可分页）
    ACADEMIC: 免费学术（限制：CURATED 数据源，最多100页）
    """
    try:
        resp = session.get(
            f"{API_BASE}entity/gene?gene_symbol=APP",
            headers={"Authorization": api_key, "Accept": "application/json"},
            timeout=10,
        )
        if resp.status_code == 200:
            profile = resp.json().get("userinfo", {}).get("profile", "UNKNOWN")
            return profile
    except Exception:
        pass
    return "UNKNOWN"


def _normalize_disease_id(disease_id: str) -> str:
    """
    标准化疾病 ID 为 DisGeNET GDA API 所需格式

    API 要求 UMLS_C 前缀格式，如 "UMLS_C0002395"
    自动转换：C0002395 → UMLS_C0002395，UMLS_C0002395 → 保持不变
    """
    did = disease_id.strip()
    # 已经是 UMLS_C 格式
    if did.startswith("UMLS_C") and len(did) > 6:
        return did
    # 纯 C 编号格式，自动加 UMLS_ 前缀
    if did.startswith("C") and len(did) == 8 and did[1:].isdigit():
        return f"UMLS_{did}"
    # 其他 ID 格式（MONDO_, EFO_, 等）原样返回
    return did


def search_disease_by_name(session: requests.Session, name: str, api_key: str) -> list[dict]:
    """根据疾病名称搜索 DisGeNET"""
    encoded = name.replace(" ", "%20")
    url = f"{API_BASE}entity/disease?disease_free_text_search_string={encoded}"
    data = _request(session, url, api_key)
    if not data:
        return []
    return data.get("payload", [])


def get_disease_genes(
    session: requests.Session,
    disease_id: str,
    api_key: str,
    database: str = "CURATED",
    min_score: float = 0.0,
    max_score: float = 1.0,
) -> tuple[list[dict], int]:
    """
    获取疾病关联的基因列表

    Returns
    -------
    (results, total_count)
    """
    # 标准化 ID 格式：确保 UMLS_C 前缀
    disease_id = _normalize_disease_id(disease_id)

    url = (
        f"{API_BASE}gda/summary"
        f"?disease={disease_id}"
        f"&source={database}"
        f"&min_score={min_score}"
        f"&max_score={max_score}"
        f"&order_by=score"
        f"&page_number=0"
    )
    data = _request(session, url, api_key)
    if not data:
        return [], 0

    total = data.get("paging", {}).get("totalElements", 0)
    results = list(data.get("payload", []))

    total_pages = min((total + 99) // 100, 100)
    logger.info(f"第 1/{total_pages} 页完成，本页 {len(results)} 条，累计 {len(results)} 条")

    for page in range(1, total_pages):
        time.sleep(PAGE_DELAY)
        page_url = url.replace("page_number=0", f"page_number={page}")
        pdata = _request(session, page_url, api_key)
        if not pdata:
            logger.warning(f"第 {page + 1} 页失败，跳过")
            continue
        page_results = pdata.get("payload", [])
        results.extend(page_results)
        logger.info(f"第 {page + 1}/{total_pages} 页完成，本页 {len(page_results)} 条，累计 {len(results)} 条")

    return results, total


def fetch_all_diseases(
    session: requests.Session,
    api_key: str,
) -> tuple[list[dict], int, int]:
    """
    分页抓取 DisGeNET 学术许可下的全部疾病实体列表。

    调用 /entity/disease 端点（不传 disease_free_text_search_string），
    自动翻页至 API 返回上限（100 页 = 10,000 条）。

    Returns
    -------
    (diseases, fetched, total)
        diseases : 获取到的疾病实体列表
        fetched  : 实际获取条数（≤10,000）
        total    : API 报告的总数（≥18,000）
    """
    url = f"{API_BASE}entity/disease?page_number=0"
    data = _request(session, url, api_key)
    if not data:
        return [], 0, 0

    total = data.get("paging", {}).get("totalElements", 0)
    results = list(data.get("payload", []))

    if total <= 0:
        return results, len(results), total

    # 学术许可上限 100 页（0~99）
    max_api_pages = min((total + 99) // 100, 100)
    logger.info(f"全量疾病：API 报告 {total} 条，分页上限 {max_api_pages} 页，开始抓取...")
    logger.info(f"第 1/{max_api_pages} 页完成，本页 {len(results)} 条，累计 {len(results)} 条")

    for page in range(1, max_api_pages):
        time.sleep(PAGE_DELAY)
        page_url = f"{API_BASE}entity/disease?page_number={page}"
        pdata = _request(session, page_url, api_key)
        if not pdata:
            logger.warning(f"第 {page + 1} 页失败，跳过")
            continue
        page_results = pdata.get("payload", [])
        results.extend(page_results)
        logger.info(f"第 {page + 1}/{max_api_pages} 页完成，本页 {len(page_results)} 条，累计 {len(results)} 条")

    return results, len(results), total


# ===================================================================
#  外部疾病本体下载（OBO 格式流式解析，免费无需 API Key）
# ===================================================================

def _parse_obo_stream(url: str, id_prefix: str, timeout: int = 300) -> list[dict]:
    """
    流式下载并解析 OBO 文件，提取符合 id_prefix 的 [Term] 块。

    OBO 格式示例::
        [Term]
        id: DOID:0001816
        name: angiosarcoma
        def: "A vascular cancer ..." [...]
        synonym: "hemangiosarcoma" EXACT []
        xref: UMLS_CUI:C0018923
        is_a: DOID:175 ! vascular cancer

    Parameters
    ----------
    url : str
        OBO 文件下载地址
    id_prefix : str
        仅提取 id 以此前缀开头的术语（如 "DOID:"、"MONDO:"）
    timeout : int
        下载超时秒数（MONDO 文件约 51 MB，需较长超时）

    Returns
    -------
    list[dict]
        每个元素含 keys: obo_id, name, definition, umls_cui, synonyms, xrefs, parents
    """
    import re

    terms: list[dict] = []
    in_term = False
    current: dict[str, list[str]] = {}
    line_count = 0
    last_log = 0

    logger.info(f"正在下载 OBO 文件: {url}")
    resp = requests.get(url, timeout=timeout, stream=True)
    resp.raise_for_status()

    # 读取总大小用于进度展示
    content_length = resp.headers.get("content-length")
    total_size = int(content_length) if content_length else None
    downloaded = 0

    leftover = ""
    for chunk in resp.iter_content(8192):
        downloaded += len(chunk)
        text = leftover + chunk.decode("utf-8", errors="replace")
        lines = text.split("\n")
        # 最后一行可能不完整，留到下次
        leftover = lines[-1]
        lines = lines[:-1]

        for line in lines:
            line_count += 1
            stripped = line.strip()

            if stripped == "[Term]":
                # 保存上一个 term
                if in_term and current:
                    term = _finalize_obo_term(current, id_prefix)
                    if term:
                        terms.append(term)
                in_term = True
                current = {}
            elif in_term:
                if stripped == "":
                    # 空行表示 term 结束
                    if current:
                        term = _finalize_obo_term(current, id_prefix)
                        if term:
                            terms.append(term)
                    in_term = False
                    current = {}
                else:
                    # 解析 key: value 行
                    if ":" in stripped:
                        key, _, value = stripped.partition(":")
                        value = value.strip()
                        current.setdefault(key, []).append(value)

        # 进度日志
        if total_size and line_count - last_log >= 500000:
            pct = downloaded / total_size * 100
            logger.info(f"  下载解析中... {pct:.0f}% ({len(terms)} 个术语已提取)")
            last_log = line_count

    # 处理最后一个 term
    if in_term and current:
        term = _finalize_obo_term(current, id_prefix)
        if term:
            terms.append(term)

    logger.info(f"OBO 解析完成：筛选出 {len(terms)} 个 {id_prefix} 术语")
    return terms


def _finalize_obo_term(raw: dict[str, list[str]], id_prefix: str) -> dict | None:
    """将原始 key→[values] 映射转为结构化 dict，仅保留含 id_prefix 的术语"""
    ids = raw.get("id", [])
    if not ids or not ids[0].startswith(id_prefix):
        return None

    def _first(key: str) -> str:
        vals = raw.get(key, [])
        return vals[0] if vals else ""

    def _join(key: str) -> str:
        return "; ".join(raw.get(key, []))

    # 提取 UMLS CUI（兼容两种格式：UMLS_CUI:C0018923 和 UMLS:C0018923）
    umls_cuis = []
    other_xrefs = []
    for x in raw.get("xref", []):
        # 去掉 OBO 的 source 标注 {source=...}
        x_clean = x.split("{")[0].strip()
        if x_clean.startswith("UMLS_CUI:"):
            umls_cuis.append(x_clean.split(":", 1)[1].strip())
        elif x_clean.startswith("UMLS:"):
            umls_cuis.append(x_clean.split(":", 1)[1].strip())
        else:
            other_xrefs.append(x)

    # 提取同义词（去掉 EXACT/RELATED 等标记）
    synonyms = []
    for s in raw.get("synonym", []):
        # 格式: "synonym text" EXACT []
        if s.startswith('"'):
            end = s.find('"', 1)
            if end > 0:
                synonyms.append(s[1:end])

    return {
        "obo_id": ids[0],
        "name": _first("name"),
        "definition": _first("def"),
        "umls_cui": "; ".join(umls_cuis),
        "synonyms": "; ".join(synonyms[:10]),  # 最多 10 个
        "xrefs": "; ".join(other_xrefs[:20]),
        "parents": "; ".join(raw.get("is_a", [])),
    }


def fetch_ontology_diseases(onto_key: str) -> tuple[list[dict], str, str]:
    """
    下载并解析指定疾病本体的 OBO 文件。

    Parameters
    ----------
    onto_key : str
        "DO" 或 "MONDO"

    Returns
    -------
    (terms, onto_name, source_url)
    """
    cfg = OBO_SOURCES.get(onto_key)
    if not cfg:
        raise ValueError(f"未知本体: {onto_key}，可选: {list(OBO_SOURCES.keys())}")

    terms = _parse_obo_stream(cfg["url"], cfg["id_prefix"])
    return terms, cfg["name"], cfg["url"]


# ===================================================================
#  数据处理与保存
# ===================================================================

DB_NAME = "DisGeNET"


def _safe_name(text: str) -> str:
    """文件名安全处理：移除特殊字符，保留字母数字与词间空格"""
    safe = "".join(c for c in text if c.isalnum() or c in " _-()（）")
    return " ".join(safe.split()) or "output"


def _output_dir() -> str:
    """输出子文件夹（与脚本同名，如 disgenet），不存在时自动创建"""
    script_name = os.path.splitext(os.path.basename(__file__))[0].lower()
    out = os.path.join(os.path.dirname(os.path.abspath(__file__)), script_name)
    os.makedirs(out, exist_ok=True)
    return out


def _xl(value):
    """将 openpyxl 无法直接写入的复杂类型（list/dict）转为字符串"""
    if isinstance(value, (list, tuple)):
        return ", ".join(str(v) for v in value)
    if isinstance(value, dict):
        return json.dumps(value, ensure_ascii=False)
    return value


def _gene_row_values(idx: int, row: dict) -> list:
    """构建单行基因-疾病关联数据（与表头顺序一致）"""
    return [
        idx,
        row.get("gene_symbol", row.get("symbolOfGene", "")),
        row.get("geneid", row.get("geneNcbiID", "")),
        row.get("geneEnsemblIDs", row.get("ensemblid", "")),
        row.get("geneProteinStrIDs", row.get("uniprotids", "")),
        row.get("geneProteinClassNames", row.get("protein_class_name", "")),
        row.get("diseaseName", row.get("disease_name", "")),
        row.get("diseaseid", ""),
        row.get("score"),
        row.get("geneDSI"),
        row.get("geneDPI"),
        row.get("genepLI"),
        row.get("ei", row.get("evidence_index")),
        row.get("numPMIDs"),
        row.get("yearInitial"),
        row.get("yearFinal"),
    ]


def save_results(rows: list[dict], disease_name: str, disease_id: str, keyword: str) -> str:
    """保存结果到 Excel"""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filepath = os.path.join(
        _output_dir(), f"{DB_NAME}_{_safe_name(keyword)}_{timestamp}.xlsx"
    )

    # 按 score 降序
    rows_sorted = sorted(
        rows,
        key=lambda r: float(r.get("score", 0) or 0),
        reverse=True,
    )

    wb = openpyxl.Workbook()

    # ---- Sheet 1: 基因-疾病关联 ----
    ws1 = wb.active
    ws1.title = "基因疾病关联"
    headers = [
        "排名", "基因 Symbol", "基因 ID (NCBI)", "Ensembl ID", "UniProt ID",
        "蛋白分类", "疾病名称", "疾病 UMLS ID", "综合评分(score)",
        "DSI", "DPI", "基因 pLI", "证据指数(EI)",
        "文献数(PMID)", "最早年份", "最新年份",
    ]
    ws1.append(headers)

    for idx, row in enumerate(rows_sorted, start=1):
        ws1.append([_xl(v) for v in _gene_row_values(idx, row)])

    # 列宽
    widths = {"A": 6, "B": 16, "C": 14, "D": 20, "E": 22, "F": 30, "G": 40, "H": 18,
              "I": 12, "J": 8, "K": 8, "L": 10, "M": 12, "N": 10, "O": 10, "P": 10}
    for col, w in widths.items():
        ws1.column_dimensions[col].width = w
    ws1.freeze_panes = "A2"

    # ---- Sheet 2: Top 100 ----
    ws2 = wb.create_sheet("TOP100高评分")
    ws2.append(headers)
    for idx, row in enumerate(rows_sorted[:100], start=1):
        ws2.append([_xl(v) for v in _gene_row_values(idx, row)])
    for col, w in widths.items():
        ws2.column_dimensions[col].width = w
    ws2.freeze_panes = "A2"

    # ---- Sheet 3: 汇总 ----
    ws3 = wb.create_sheet("汇总")
    ws3.append(["项目", "内容"])
    ws3.append(["数据库", "DisGeNET"])
    ws3.append(["API", "http://api.disgenet.com/api/v1/"])
    ws3.append(["疾病名称", disease_name])
    ws3.append(["疾病 ID", disease_id])
    ws3.append(["搜索关键词", keyword])
    ws3.append(["关联基因总数", len(rows)])
    scores = [float(r.get("score", 0) or 0) for r in rows]
    if scores:
        ws3.append(["最高评分", f"{max(scores):.2f}"])
        ws3.append(["平均评分", f"{sum(scores)/len(scores):.2f}"])
        high = sum(1 for s in scores if s >= 0.7)
        mid = sum(1 for s in scores if 0.3 <= s < 0.7)
        ws3.append(["高评分(≥0.7)基因数", high])
        ws3.append(["中评分(0.3-0.7)基因数", mid])
    ws3.append(["生成时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    ws3.column_dimensions["A"].width = 22
    ws3.column_dimensions["B"].width = 55

    wb.save(filepath)
    logger.info(f"文件已保存: {filepath}")
    return filepath


def export_disease_hits(hits: list[dict], keyword: str) -> str:
    """
    将疾病名称检索结果（匹配到的疾病实体列表）导出为 Excel 文件

    Parameters
    ----------
    hits : list[dict]
        search_disease_by_name 返回的疾病实体列表（完整结果）
    keyword : str
        用户搜索关键词（用于文件命名）

    Returns
    -------
    str
        导出文件的完整路径
    """
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filepath = os.path.join(
        _output_dir(), f"{DB_NAME}_{_safe_name(keyword)}_疾病检索_{timestamp}.xlsx"
    )

    # 优先展示的常见字段，其余字段动态追加，确保完整导出
    preferred = [
        "name", "diseaseUMLSCUI", "diseaseType",
        "numPMIDs", "numGenes", "numVariants",
        "diseaseClasses_MSH", "diseaseClasses_UMLS_ST",
        "diseaseClasses_DO", "diseaseClasses_HPO",
    ]
    # 按 preferred 顺序在前，再追加其余实际出现过的字段
    ordered_keys = list(preferred)
    for hit in hits:
        for k in hit.keys():
            if k not in ordered_keys:
                ordered_keys.append(k)
    # 仅保留真实存在于数据中的字段，避免空列
    present_keys = [k for k in ordered_keys if any(k in hit for hit in hits)]

    def _cell(value):
        """将 list/dict 等复杂类型转为可读字符串"""
        if isinstance(value, (list, tuple)):
            return ", ".join(str(v) for v in value)
        if isinstance(value, dict):
            return json.dumps(value, ensure_ascii=False)
        return value

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "疾病检索结果"
    ws.append(["序号"] + present_keys)
    for idx, hit in enumerate(hits, start=1):
        ws.append([idx] + [_cell(hit.get(k, "")) for k in present_keys])
    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 6
    for i in range(len(present_keys)):
        ws.column_dimensions[get_column_letter(i + 2)].width = 22
    wb.save(filepath)

    logger.info(f"疾病检索结果已导出: {filepath}")
    return filepath


def _prompt_export_hits(hits: list[dict], keyword: str):
    """询问用户是否将完整疾病检索结果导出为 Excel 文件（可选操作）"""
    ans = input("\n  是否将完整检索结果导出为 Excel 文件？(y/N): ").strip().lower()
    if ans not in ("y", "yes"):
        return
    try:
        path = export_disease_hits(hits, keyword)
        print(f"  ✓ 已导出 {len(hits)} 条结果: {path}")
    except Exception as e:
        logger.error(f"导出失败: {e}")
        print(f"  ✗ 导出失败: {e}")


def save_disease_list(diseases: list[dict]) -> str:
    """
    将全量疾病实体列表导出为 Excel 文件

    字段按常用优先排列（name, diseaseUMLSCUI 在前），
    其余字段动态追加，确保不丢列。
    """
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filepath = os.path.join(
        _output_dir(), f"{DB_NAME}_全部疾病列表_{timestamp}.xlsx"
    )

    preferred = [
        "name", "diseaseUMLSCUI", "type",
        "diseaseClasses_MSH", "diseaseClasses_UMLS_ST",
        "diseaseClasses_DO", "diseaseClasses_HPO",
        "disease_prevalence_class", "disease_inheritance",
        "numGenesAssociatedToDisease", "numVariantsAssociatedToDisease",
        "numPublications", "search_rank", "synonyms",
    ]
    ordered_keys = list(preferred)
    for d in diseases:
        for k in d.keys():
            if k not in ordered_keys:
                ordered_keys.append(k)
    present_keys = [k for k in ordered_keys if any(k in d for d in diseases)]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "全部疾病列表"
    ws.append(["序号"] + present_keys)
    for idx, d in enumerate(diseases, start=1):
        ws.append([idx] + [_xl(d.get(k, "")) for k in present_keys])
    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 6
    col_widths = {"name": 50, "diseaseUMLSCUI": 16, "diseaseClasses_MSH": 55,
                  "diseaseClasses_UMLS_ST": 40, "diseaseClasses_DO": 45,
                  "diseaseClasses_HPO": 45, "synonyms": 50}
    for i, key in enumerate(present_keys):
        ws.column_dimensions[get_column_letter(i + 2)].width = col_widths.get(key, 22)
    wb.save(filepath)
    logger.info(f"疾病列表已保存: {filepath}")
    return filepath


def save_ontology_diseases(terms: list[dict], onto_name: str, onto_key: str) -> str:
    """
    将疾病本体术语导出为 Excel 文件

    字段固定：序号, ID, 疾病名称, 定义, UMLS_CUI, 同义词, 其他交叉引用, 父类
    """
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_label = onto_key  # DO / MONDO
    filepath = os.path.join(
        _output_dir(), f"{safe_label}_疾病本体_{timestamp}.xlsx"
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{onto_key}疾病本体"
    headers = ["序号", "ID", "疾病名称", "定义", "UMLS CUI", "同义词", "其他交叉引用", "父类(is_a)"]
    ws.append(headers)

    for idx, t in enumerate(terms, start=1):
        ws.append([
            idx,
            t.get("obo_id", ""),
            t.get("name", ""),
            _xl(t.get("definition", "")),
            t.get("umls_cui", ""),
            t.get("synonyms", ""),
            t.get("xrefs", ""),
            t.get("parents", ""),
        ])

    widths = {"A": 6, "B": 18, "C": 55, "D": 80, "E": 25, "F": 60, "G": 60, "H": 60}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"
    wb.save(filepath)

    # 统计含 UMLS 映射的术语数
    umls_count = sum(1 for t in terms if t.get("umls_cui"))
    logger.info(f"{onto_name} 已保存 ({len(terms)} 术语, {umls_count} 含 UMLS CUI): {filepath}")
    return filepath


# ===================================================================
#  交互菜单
# ===================================================================

def clear_screen():
    os.system("cls" if os.name == "nt" else "clear")


def press_enter():
    input("\n按 Enter 返回主菜单...")


def show_banner():
    clear_screen()
    print("╔" + "═" * 58 + "╗")
    print("║" + "   DisGeNET 疾病靶点检索工具 v1.4".center(56) + "║")
    print("║" + "   数据源: https://www.disgenet.com/".center(48) + "║")
    print("║" + "   需先申请免费学术许可证获取 API Key".center(44) + "║")
    print("║" + "   author: shenjianlin".center(48) + "║")
    print("║" + "   site: git@github.com:sjl-openmywork/Network-toxicology-script.git".center(38) + "║")
    print("╚" + "═" * 58 + "╝")


def _show_preset():
    print("  ┌────────────────────────────────────────────────┐")
    print("  │  快捷疾病 UMLS ID 列表                           │")
    for key, info in PRESET_DISEASES.items():
        print(f"  │  {key:>2}. {info['name']:<42} │")
    print("  └────────────────────────────────────────────────┘")


def menu_search_by_name(session: requests.Session, api_key: str):
    """菜单 [1] 按疾病名称搜索"""
    show_banner()
    print("\n  ▸ 按疾病名称搜索基因")
    print("  " + "─" * 54)
    print("  输入疾病英文名称，自动搜索匹配的疾病实体\n")
    _show_preset()
    print()

    name = input("  请输入疾病英文名称或快捷编号: ").strip()
    if not name:
        press_enter()
        return

    # 快捷编号或名称匹配
    disease_id = ""
    disease_name = ""

    if name in PRESET_DISEASES:
        info = PRESET_DISEASES[name]
        disease_id = info["id"]
        disease_name = info["name"]
        print(f"\n  已匹配预置编号: {info['name']} ({info['id']})")
    else:
        matched = None
        for info in PRESET_DISEASES.values():
            if info["name"].lower() == name.lower():
                matched = info
                break
        if matched:
            disease_id = matched["id"]
            disease_name = matched["name"]
            print(f"\n  已匹配预置名称: {matched['name']} ({matched['id']})")

    # 未匹配预置 → API 搜索
    if not disease_id:
        print(f"\n  正在搜索「{name}」...")
        hits = search_disease_by_name(session, name, api_key)
        if not hits:
            print(f"  ✗ 未找到「{name}」，请尝试其他关键词")
            press_enter()
            return

        print(f"\n  找到 {len(hits)} 个匹配疾病：")
        # 展示全部匹配结果（不再限制前 10 项）
        for i, hit in enumerate(hits, start=1):
            cui = hit.get("diseaseUMLSCUI", "N/A")
            print(f"  [{i}] {hit.get('name', '')}  (UMLS: {cui})")

        # 可选：将完整检索结果导出为表格文件（Excel/CSV）
        _prompt_export_hits(hits, name)

        while True:
            choice = input(f"\n  请选择 [1-{len(hits)}] 或 0 取消: ").strip()
            if choice == "0":
                press_enter()
                return
            try:
                idx = int(choice) - 1
                if 0 <= idx < len(hits):
                    hit = hits[idx]
                    disease_name = hit.get("name", name)
                    disease_id = hit.get("diseaseUMLSCUI", "")
                    if not disease_id:
                        print("  ⚠ 该疾病无 UMLS ID，无法查询 GDA")
                        press_enter()
                        return
                    break
            except ValueError:
                pass
            print("  ⚠ 输入有误")

    _do_search(session, api_key, disease_id, disease_name, name)


def menu_search_by_id(session: requests.Session, api_key: str):
    """菜单 [2] 直接输入疾病 ID"""
    show_banner()
    print("\n  ▸ 直接输入疾病 ID")
    print("  " + "─" * 54)
    print("  支持: UMLS (C0028754), MONDO, OMIM, MeSH, DO, EFO, ORDO, HPO\n")
    _show_preset()
    print()

    raw = input("  请输入疾病 ID（或快捷编号）: ").strip()
    if not raw:
        press_enter()
        return

    if raw in PRESET_DISEASES:
        info = PRESET_DISEASES[raw]
        disease_id = info["id"]
        disease_name = info["name"]
    elif raw.startswith(("C", "MONDO_", "OMIM_", "MESH_", "EFO_", "ORDO_", "DOID_", "HP_")):
        disease_id = raw
        disease_name = raw
    else:
        # 尝试作为 UMLS ID
        disease_id = f"C{raw}" if not raw.startswith("C") else raw
        disease_name = disease_id

    _do_search(session, api_key, disease_id, disease_name, disease_name)


def _do_search(session: requests.Session, api_key: str, disease_id: str, disease_name: str, keyword: str):
    """执行搜索并保存"""
    print(f"\n  疾病: {disease_name}  ID: {disease_id}")

    # 数据库选择
    print(f"\n  可用数据源: {', '.join(ACADEMIC_SOURCES)}")
    db = input(f"  选择数据源（默认 CURATED）: ").strip().upper() or "CURATED"
    if db not in ACADEMIC_SOURCES:
        print(f"  ⚠ 学术计划不支持「{db}」，已使用 CURATED")
        db = "CURATED"

    # 评分过滤
    try:
        min_s = input("  最低评分过滤（0-1，默认 0）: ").strip()
        min_score = float(min_s) if min_s else 0.0
        max_s = input("  最高评分过滤（0-1，默认 1）: ").strip()
        max_score = float(max_s) if max_s else 1.0
    except ValueError:
        print("  ⚠ 输入无效，使用默认范围 0-1")
        min_score, max_score = 0.0, 1.0

    print()
    logger.info(f"正在检索疾病「{disease_name}」的关联基因 (source={db}, score={min_score}-{max_score})...")

    rows, total = get_disease_genes(session, disease_id, api_key, db, min_score, max_score)

    if not rows:
        print(f"\n  ✗ 未找到关联基因")
        print(f"  可能原因：疾病 ID 无效 / 数据库中无此疾病 / 评分范围过窄")
        press_enter()
        return

    # 保存
    filepath = save_results(rows, disease_name, disease_id, keyword)

    # 统计
    scores = [float(r.get("score", 0) or 0) for r in rows]
    print(f"\n  ✓ 疾病: {disease_name}")
    print(f"  ✓ 关联基因总数: {total}  |  已获取: {len(rows)} 条")
    if scores:
        high = sum(1 for s in scores if s >= 0.7)
        print(f"  ✓ 高评分(≥0.7): {high} 个  |  最高评分: {max(scores):.2f}")
    print(f"  ✓ 文件: {filepath}")
    print(f"  ✓ Excel 含 3 个 Sheet：基因疾病关联 | TOP100高评分 | 汇总")

    # 预览 Top 10
    rows_sorted = sorted(rows, key=lambda r: float(r.get("score", 0) or 0), reverse=True)
    print(f"\n  TOP 10 基因预览：")
    print(f"  {'排名':<5} {'基因Symbol':<16} {'疾病':<35} {'评分':<8} {'DSI':<7} {'DPI':<7} {'PMID':<6}")
    print(f"  {'-'*84}")
    for i, r in enumerate(rows_sorted[:10], start=1):
        print(f"  {i:<5} {r.get('gene_symbol', r.get('symbolOfGene', '')):<16} "
              f"{(r.get('diseaseName', r.get('disease_name', '')) or '')[:33]:<35} "
              f"{float(r.get('score', 0) or 0):.4f}  "
              f"{r.get('geneDSI', ''):<7} {r.get('geneDPI', ''):<7} {r.get('numPMIDs', ''):<6}")

    press_enter()


def menu_fetch_all_diseases(session: requests.Session, api_key: str):
    """菜单 [3] 下载全部疾病列表"""
    show_banner()
    print("\n  ▸ 下载全部疾病列表")
    print("  " + "─" * 54)
    print("  通过 entity/disease 端点（不传搜索词）分页抓取 CURATED 全量疾病元数据\n")
    print("  学术许可上限 100 页 × 100 条/页 = 最多 10,000 条\n")

    confirm = input("  预计耗时约 150 秒（100 页 × 1.5s），是否继续？(y/N): ").strip().lower()
    if confirm not in ("y", "yes"):
        print("  已取消")
        press_enter()
        return

    print()
    logger.info("开始分页抓取全量疾病列表...")
    diseases, fetched, total = fetch_all_diseases(session, api_key)

    if not diseases:
        print("\n  ✗ 未获取到疾病数据")
        press_enter()
        return

    filepath = save_disease_list(diseases)
    print(f"\n  ✓ API 报告疾病总数: {total}")
    print(f"  ✓ 实际抓取: {fetched} 条（学术许可上限 10,000 条）")
    coverage = f"{fetched / total * 100:.1f}%" if total > 0 else "N/A"
    print(f"  ✓ 覆盖率: {coverage}")
    print(f"  ✓ 文件: {filepath}")

    # 预览
    print(f"\n  前 5 个疾病预览：")
    for i, d in enumerate(diseases[:5], start=1):
        print(f"    {i}. {d.get('name', '')}  (UMLS: {d.get('diseaseUMLSCUI', 'N/A')})")
    print(f"    ... 共 {fetched} 条")

    press_enter()


def menu_fetch_ontology(onto_key: str):
    """菜单 [4][5] 下载疾病本体（DO / MONDO）"""
    cfg = OBO_SOURCES[onto_key]
    show_banner()
    print(f"\n  ▸ 下载 {cfg['name']}")
    print("  " + "─" * 54)
    print(f"  从官方 OBO 文件流式解析疾病术语及 UMLS CUI 交叉映射")
    print(f"  数据源: {cfg['url']}")

    if onto_key == "MONDO":
        print("\n  ⚠ MONDO OBO 文件约 51 MB，下载+解析预计 30~60 秒")

    confirm = input("\n  是否继续？(y/N): ").strip().lower()
    if confirm not in ("y", "yes"):
        print("  已取消")
        press_enter()
        return

    print()
    logger.info(f"开始下载 {cfg['name']} ...")
    try:
        terms, onto_name, source_url = fetch_ontology_diseases(onto_key)
    except Exception as e:
        logger.error(f"下载失败: {e}")
        print(f"\n  ✗ 下载失败: {e}")
        press_enter()
        return

    if not terms:
        print(f"\n  ✗ 未提取到 {cfg['id_prefix']} 术语")
        press_enter()
        return

    filepath = save_ontology_diseases(terms, onto_name, onto_key)
    umls_count = sum(1 for t in terms if t.get("umls_cui"))
    print(f"\n  ✓ {onto_name}")
    print(f"  ✓ 共 {len(terms)} 个疾病术语")
    print(f"  ✓ 其中 {umls_count} 个含 UMLS CUI 交叉映射（可直接对接 DisGeNET）")
    print(f"  ✓ 文件: {filepath}")
    print(f"\n  前 5 个术语预览：")
    for i, t in enumerate(terms[:5], start=1):
        cui = t.get("umls_cui", "")[:30] or "N/A"
        print(f"    {i}. [{t.get('obo_id','')}] {t.get('name','')[:55]}")
        if cui != "N/A":
            print(f"       UMLS: {cui}")

    press_enter()


# ===================================================================
#  主入口
# ===================================================================

def main():
    # 获取 API Key
    api_key = get_api_key()
    if not api_key:
        print()
        print("━" * 58)
        print("  如需使用 DisGeNET 脚本，请先申请 API Key：")
        print()
        print("  申请流程：")
        print("    1. 访问 https://disgenet.com/academic-apply")
        print("    2. 用机构邮箱注册.edu/.edu.cn")
        print("    3. 审核通过后获取 API Key（通常 2-7 天）")
        print("    4. 运行本脚本输入 Key")
        print()
        print("  已有 Key 的话请重新运行，或设置环境变量：")
        print("    set DISGENET_API_KEY=your_key_here")
        print("━" * 58)
        press_enter()
        return

    session = requests.Session()

    # 检测账户类型
    print("\n  正在验证 API Key...")
    profile = check_account_profile(session, api_key)
    if profile == "ACADEMIC":
        print("  ✓ 账户类型: ACADEMIC（正常）")
        time.sleep(0.5)
    elif profile == "TRIAL":
        print()
        print("  ⚠ 当前账户为 TRIAL（7天试用），有以下限制：")
        print("    • 仅返回前 30 条结果，不可分页")
        print("    • 学术申请审核通过后将自动切换为 ACADEMIC")
        print("    • 如已完成申请确认但仍是 TRIAL，请等待审核（2-7天）")
        time.sleep(2)
    else:
        print(f"  ⚠ 无法确定账户类型: {profile}")
        time.sleep(1)

    while True:
        show_banner()
        print()
        print("  ┌────────────────────────────────────────────────────┐")
        print("  │  [1] 按疾病名称搜索        输入名称 → 选择 → 基因 │")
        print("  │  [2] 直接输入疾病 ID       UMLS/MONDO/OMIM 等    │")
        print("  │  [3] 下载DisGeNET疾病列表  分页抓取全量疾病元数据  │")
        print("  │  [4] 下载Disease Ontology   DO疾病本体+UMLS映射   │")
        print("  │  [5] 下载MONDO本体          MONDO疾病本体+UMLS映射 │")
        print("  │  [0] 退出                                          │")
        print("  └────────────────────────────────────────────────────┘")
        print()
        print("  提示: 支持快捷编号（如 1=Alzheimer, 2=Diabetes）")

        choice = input("\n  请选择 [0-5]: ").strip()

        if choice == "1":
            menu_search_by_name(session, api_key)
        elif choice == "2":
            menu_search_by_id(session, api_key)
        elif choice == "3":
            menu_fetch_all_diseases(session, api_key)
        elif choice == "4":
            menu_fetch_ontology("DO")
        elif choice == "5":
            menu_fetch_ontology("MONDO")
        elif choice == "0":
            show_banner()
            print("\n  感谢使用 DisGeNET 疾病靶点检索工具，再见！\n")
            break
        else:
            print(f"\n  ⚠ 无效选项「{choice}」，请重新选择")
            time.sleep(1)


if __name__ == "__main__":
    main()
