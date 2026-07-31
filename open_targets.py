"""
Open Targets Platform 疾病靶点检索脚本 v1.0
================================================================================
功能：根据疾病名称，从 Open Targets Platform 检索关联靶点及证据评分
数据源：https://platform.opentargets.org/
API: https://api.platform.opentargets.org/api/v4/graphql (GraphQL)

=== 核心功能 ===
  - 疾病名称搜索 → 匹配疾病实体（MONDO/EFO ID）
  - 获取疾病关联的所有靶点，含整体评分 + 各数据类型评分
  - 数据类型：遗传关联、体细胞突变、RNA表达、通路、已知药物、动物模型、文献

=== 交互菜单 ===
  1. 搜索疾病并检索靶点 —— 输入疾病名 → 选择匹配项 → 获取靶点
  2. 多疾病批量查询 —— 一次输入多个疾病名，逐个检索
  0. 退出

=== 评分说明 ===
  - overall_score: 综合关联评分 (0-1)，越高越可信
  - 子评分：genetic_association, somatic_mutation, rna_expression,
            affected_pathway, known_drug, animal_model, literature
"""

import os
import sys
import time
import logging
from datetime import datetime

import openpyxl
import requests

# ============ 日志配置 ============
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger(__name__)

# ============ 常量配置 ============
API_URL = "https://api.platform.opentargets.org/api/v4/graphql"
PAGE_SIZE = 500           # 每页靶点数
REQUEST_TIMEOUT = 60      # GraphQL 请求超时（秒）
RATE_LIMIT_DELAY = 0.3    # 请求间隔

# 预置常见疾病 ID（无需搜索直接查）
PRESET_DISEASES = {
    "1":  {"id": "MONDO_0004975", "name": "Alzheimer disease"},
    "2":  {"id": "EFO_0000400",  "name": "Diabetes mellitus"},
    "3":  {"id": "MONDO_0005148", "name": "Type 2 diabetes mellitus"},
    "4":  {"id": "MONDO_0005147", "name": "Type 1 diabetes mellitus"},
    "5":  {"id": "MONDO_0005180", "name": "Parkinson disease"},
    "6":  {"id": "MONDO_0004976", "name": "Amyotrophic lateral sclerosis"},
    "7":  {"id": "MONDO_0005010", "name": "Coronary artery disease"},
    "8":  {"id": "MONDO_0004992", "name": "Breast cancer"},
    "9":  {"id": "MONDO_0004979", "name": "Asthma"},
    "10": {"id": "MONDO_0005301", "name": "Multiple sclerosis"},
    "11": {"id": "MONDO_0005105", "name": "Melanoma"},
    "12": {"id": "MONDO_0005027", "name": "Epilepsy"},
}

DATA_TYPE_LABELS = {
    "genetic_association":   "遗传关联",
    "genetic_literature":    "遗传文献",
    "somatic_mutation":      "体细胞突变",
    "rna_expression":        "RNA表达",
    "affected_pathway":      "通路影响",
    "known_drug":            "已知药物",
    "animal_model":          "动物模型",
    "literature":            "文献证据",
    "clinical":              "临床证据",
}


# ===================================================================
#  GraphQL API
# ===================================================================

def _gql(query: str, variables: dict) -> dict | None:
    """执行 GraphQL 请求"""
    try:
        resp = requests.post(
            API_URL,
            json={"query": query, "variables": variables},
            timeout=REQUEST_TIMEOUT,
            headers={"Content-Type": "application/json"},
        )
        resp.raise_for_status()
        data = resp.json()
        if "errors" in data:
            logger.error(f"GraphQL 错误: {data['errors']}")
            return None
        return data.get("data")
    except requests.exceptions.Timeout:
        logger.error("GraphQL 请求超时")
        return None
    except requests.exceptions.ConnectionError:
        logger.error("连接 Open Targets API 失败，请检查网络")
        return None
    except Exception as e:
        logger.error(f"GraphQL 请求异常: {type(e).__name__}: {e}")
        return None


def search_disease(name: str, limit: int = 10) -> list[dict]:
    """按名称搜索疾病"""
    query = """
    query SearchDisease($q: String!, $size: Int!) {
      search(queryString: $q, entityNames: ["disease"], page: {index: 0, size: $size}) {
        hits {
          id
          name
          description
        }
      }
    }
    """
    data = _gql(query, {"q": name, "size": limit})
    if not data:
        return []
    return data.get("search", {}).get("hits", [])


def fetch_targets_page(disease_id: str, index: int, size: int) -> tuple[list[dict], int]:
    """获取一页靶点关联数据，返回 (rows, total_count)"""
    query = """
    query DiseaseTargets($efoId: String!, $index: Int!, $size: Int!) {
      disease(efoId: $efoId) {
        name
        associatedTargets(page: {index: $index, size: $size}) {
          count
          rows {
            target {
              id
              approvedSymbol
              approvedName
            }
            score
            datatypeScores {
              id
              score
            }
          }
        }
      }
    }
    """
    data = _gql(query, {"efoId": disease_id, "index": index, "size": size})
    if not data or not data.get("disease"):
        return [], 0
    at = data["disease"]["associatedTargets"]
    return at["rows"], at["count"]


def fetch_all_targets(disease_id: str) -> tuple[list[dict], str]:
    """分页获取疾病的所有关联靶点，返回 (rows, disease_name)"""
    # 先获取第一页，确定总数
    rows, total = fetch_targets_page(disease_id, 0, 1)
    if not rows and total == 0:
        return [], ""

    # 获取第一页实际数据
    rows, total = fetch_targets_page(disease_id, 0, PAGE_SIZE)
    all_rows = list(rows)

    # 还需获取疾病名
    query = """
    query DiseaseName($efoId: String!) {
      disease(efoId: $efoId) { name }
    }
    """
    name_data = _gql(query, {"efoId": disease_id})
    disease_name = ""
    if name_data and name_data.get("disease"):
        disease_name = name_data["disease"]["name"]

    total_pages = (total + PAGE_SIZE - 1) // PAGE_SIZE
    logger.info(f"第 1/{total_pages} 页完成，本页 {len(rows)} 条，累计 {len(all_rows)} 条")

    for page_num in range(2, total_pages + 1):
        time.sleep(RATE_LIMIT_DELAY)
        rows, _ = fetch_targets_page(disease_id, (page_num - 1) * PAGE_SIZE, PAGE_SIZE)
        if not rows:
            logger.warning(f"第 {page_num}/{total_pages} 页为空，跳过")
            continue
        all_rows.extend(rows)
        logger.info(f"第 {page_num}/{total_pages} 页完成，本页 {len(rows)} 条，累计 {len(all_rows)} 条")

    return all_rows, disease_name


# ===================================================================
#  数据处理与保存
# ===================================================================

DB_NAME = "OpenTargets"


def _safe_name(text: str) -> str:
    """文件名安全处理：移除特殊字符，保留字母数字与词间空格"""
    safe = "".join(c for c in text if c.isalnum() or c in " _-()（）")
    return " ".join(safe.split()) or "output"


def _output_dir() -> str:
    """输出子文件夹（与脚本同名，如 open_targets），不存在时自动创建"""
    script_name = os.path.splitext(os.path.basename(__file__))[0].lower()
    out = os.path.join(os.path.dirname(os.path.abspath(__file__)), script_name)
    os.makedirs(out, exist_ok=True)
    return out


def _parse_scores(row: dict) -> dict[str, float]:
    return {ds["id"]: ds["score"] for ds in (row.get("datatypeScores") or [])}


def _collect_data_types(rows: list[dict]) -> list[str]:
    """从 API 返回数据中收集所有出现的数据类型 ID"""
    seen: dict[str, bool] = {}
    for row in rows:
        for ds in (row.get("datatypeScores") or []):
            dt_id = ds.get("id", "")
            if dt_id and dt_id not in seen:
                seen[dt_id] = True
    # 保持 DATA_TYPE_LABELS 中出现过的排在前面，其余按出现顺序
    priority = [k for k in DATA_TYPE_LABELS if k in seen]
    others = [k for k in seen if k not in priority]
    return priority + others


def save_results(rows: list[dict], disease_name: str, keyword: str) -> str:
    """保存结果到 Excel"""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filepath = os.path.join(
        _output_dir(), f"{DB_NAME}_{_safe_name(keyword)}_{timestamp}.xlsx"
    )

    # 按综合评分降序排序
    rows_sorted = sorted(rows, key=lambda r: r.get("score", 0) or 0, reverse=True)

    # 动态收集数据类型
    data_types = _collect_data_types(rows_sorted)

    wb = openpyxl.Workbook()

    # ---- Sheet 1: 全部靶点（按评分排序） ----
    ws1 = wb.active
    ws1.title = "靶点关联评分"
    headers = ["排名", "Ensembl ID", "基因符号", "基因名称", "综合评分"]
    for dt_id in data_types:
        headers.append(DATA_TYPE_LABELS.get(dt_id, dt_id))
    ws1.append(headers)

    for idx, row in enumerate(rows_sorted, start=1):
        t = row.get("target") or {}
        scores = _parse_scores(row)
        values = [
            idx, t.get("id", ""), t.get("approvedSymbol", ""),
            t.get("approvedName", ""), row.get("score"),
        ]
        for dt_id in data_types:
            values.append(scores.get(dt_id, ""))
        ws1.append(values)

    # 列宽
    widths = {"A": 6, "B": 20, "C": 16, "D": 50, "E": 10}
    for i in range(len(data_types)):
        widths[openpyxl.utils.get_column_letter(6 + i)] = 10
    for col, w in widths.items():
        ws1.column_dimensions[col].width = w
    ws1.freeze_panes = "A2"

    # ---- Sheet 2: Top 100 高评分靶点 ----
    ws2 = wb.create_sheet("TOP100高评分")
    ws2.append(headers)
    for idx, row in enumerate(rows_sorted[:100], start=1):
        t = row.get("target") or {}
        scores = _parse_scores(row)
        values = [
            idx, t.get("id", ""), t.get("approvedSymbol", ""),
            t.get("approvedName", ""), row.get("score"),
        ]
        for dt_id in data_types:
            values.append(scores.get(dt_id, ""))
        ws2.append(values)
    for col, w in widths.items():
        ws2.column_dimensions[col].width = w
    ws2.freeze_panes = "A2"

    # ---- Sheet 3: 汇总信息 ----
    ws3 = wb.create_sheet("汇总")
    ws3.append(["项目", "内容"])
    ws3.append(["数据库", "Open Targets Platform"])
    ws3.append(["API", "https://api.platform.opentargets.org/api/v4/graphql"])
    ws3.append(["疾病名称", disease_name])
    ws3.append(["搜索关键词", keyword])
    ws3.append(["关联靶点总数", len(rows)])
    ws3.append(["证据数据类型"] + [DATA_TYPE_LABELS.get(dt, dt) for dt in data_types])
    # 分数分布
    valid_scores = [r.get("score") for r in rows if r.get("score") is not None]
    if valid_scores:
        ws3.append(["最高评分", f"{max(valid_scores):.4f}"])
        ws3.append(["平均评分", f"{sum(valid_scores)/len(valid_scores):.4f}"])
        high = sum(1 for s in valid_scores if s >= 0.7)
        mid = sum(1 for s in valid_scores if 0.3 <= s < 0.7)
        low = sum(1 for s in valid_scores if s < 0.3)
        ws3.append(["高评分(≥0.7)靶点数", high])
        ws3.append(["中评分(0.3-0.7)靶点数", mid])
        ws3.append(["低评分(<0.3)靶点数", low])
    ws3.append(["生成时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    ws3.column_dimensions["A"].width = 22
    ws3.column_dimensions["B"].width = 55

    wb.save(filepath)
    logger.info(f"文件已保存: {filepath}")
    return filepath


# ===================================================================
#  交互菜单
# ===================================================================

def clear_screen():
    os.system("cls" if os.name == "nt" else "clear")


def press_enter():
    input("\n按 Enter 返回主菜单...")


def show_banner():
    clear_screen()
    print("╔" + "═" * 60 + "╗")
    print("║" + "   Open Targets Platform 疾病靶点检索 v1.0".center(58) + "║")
    print("║" + "   数据源: https://platform.opentargets.org/".center(49) + "║")
    print("║" + "   author: shenjianlin".center(50) + "║")
    print("║" + "   site: git@github.com:sjl-openmywork/Network-toxicology-script.git".center(40) + "║")
    print("╚" + "═" * 60 + "╝")


def _show_preset_diseases():
    print("  ┌──────────────────────────────────────────────┐")
    print("  │  快捷疾病 ID 列表                              │")
    for key, info in PRESET_DISEASES.items():
        print(f"  │  {key:>2}. {info['name']:<40} {info['id']} │")
    print("  └──────────────────────────────────────────────┘")


def _pick_disease() -> tuple[str | None, str, str]:
    """
    让用户选择疾病：输入名称搜索 或 输入预置编号 或 直接输ID

    Returns
    -------
    (disease_id, disease_name, keyword) or (None, "", "") if cancelled
    """
    keyword = input("  请输入疾病名称（支持中英文）或快捷编号: ").strip()
    if not keyword:
        print("\n  ⚠ 输入不能为空")
        return None, "", ""

    # 检查是否预置编号
    if keyword in PRESET_DISEASES:
        info = PRESET_DISEASES[keyword]
        return info["id"], info["name"], info["name"]

    # 检查是否直接输入了疾病 ID（MONDO_xxx / EFO_xxx 等）
    if keyword.startswith(("MONDO_", "EFO_", "Orphanet_", "HP_")):
        return keyword, keyword, keyword

    # 搜索
    print(f"\n  正在搜索「{keyword}」...")
    hits = search_disease(keyword)
    if not hits:
        print(f"  ✗ 未找到「{keyword}」相关疾病")
        return None, "", ""

    print(f"\n  找到 {len(hits)} 个匹配疾病：")
    for i, hit in enumerate(hits, start=1):
        desc = (hit.get("description") or "")[:80]
        print(f"  [{i}] {hit['id']}: {hit['name']}")
        if desc:
            print(f"      {desc}")

    while True:
        choice = input(f"\n  请选择 [1-{len(hits)}] 或输入 0 取消: ").strip()
        if choice == "0":
            return None, "", ""
        try:
            idx = int(choice) - 1
            if 0 <= idx < len(hits):
                hit = hits[idx]
                return hit["id"], hit["name"], keyword
        except ValueError:
            pass
        print(f"  ⚠ 请输入 1-{len(hits)} 或 0")


def menu_search():
    """菜单 [1] 搜索疾病并检索靶点"""
    show_banner()
    print("\n  ▸ 搜索疾病并检索靶点")
    print("  " + "─" * 56)
    _show_preset_diseases()
    print()

    disease_id, disease_name, keyword = _pick_disease()
    if not disease_id:
        press_enter()
        return

    print(f"\n  已选择: {disease_name} ({disease_id})")

    # 数量确认
    total = _gql("""
    query TotalCount($efoId: String!) {
      disease(efoId: $efoId) {
        associatedTargets(page: {index: 0, size: 0}) { count }
      }
    }
    """, {"efoId": disease_id})

    total_count = 0
    if total and total.get("disease"):
        total_count = total["disease"]["associatedTargets"]["count"]

    logger.info(f"疾病「{disease_name}」共有 {total_count} 个关联靶点")

    if total_count > 1000:
        print(f"  ⚠ 该疾病有 {total_count} 个靶点，获取全部预计耗时 {total_count/PAGE_SIZE*0.3:.0f} 秒")
        while True:
            limit_choice = input("  获取全部 (A) / 仅获取前 1000 (1k) / 取消 (C): ").strip().lower()
            if limit_choice == "c":
                press_enter()
                return
            elif limit_choice == "1k":
                max_count = 1000
                break
            elif limit_choice == "a":
                max_count = total_count
                break
            print("  ⚠ 请输入 A / 1k / C")
    else:
        max_count = total_count

    print()
    logger.info(f"正在获取靶点数据（每页 {PAGE_SIZE} 条）...")

    # 分页获取
    all_rows, dn = fetch_all_targets(disease_id)
    if not all_rows:
        print(f"\n  ✗ 未获取到靶点数据")
        press_enter()
        return

    # 如果用户限制数量
    if max_count < len(all_rows):
        all_rows = all_rows[:max_count]
        logger.info(f"已限制为前 {max_count} 条")

    # 按评分排序
    all_rows.sort(key=lambda r: r.get("score") or 0, reverse=True)

    # 保存
    filepath = save_results(all_rows, disease_name or dn, keyword)

    # 统计
    valid_scores = [r.get("score") for r in all_rows if r.get("score") is not None]
    print(f"\n  ✓ 疾病: {disease_name or dn}")
    print(f"  ✓ 关联靶点: {len(all_rows)} 个")
    if valid_scores:
        high = sum(1 for s in valid_scores if s >= 0.7)
        print(f"  ✓ 高评分(≥0.7): {high} 个  |  最高评分: {max(valid_scores):.4f}")
    print(f"  ✓ 文件: {filepath}")
    print(f"  ✓ Excel 含 3 个 Sheet：靶点关联评分 | TOP100高评分 | 汇总信息")

    # 预览 Top 10
    print(f"\n  TOP 10 靶点预览：")
    print(f"  {'排名':<5} {'基因':<14} {'靶点名称':<45} {'综合评分':<8}")
    print(f"  {'-'*72}")
    for i, row in enumerate(all_rows[:10], start=1):
        t = row.get("target") or {}
        print(f"  {i:<5} {t.get('approvedSymbol', ''):<14} "
              f"{t.get('approvedName', '')[:42]:<45} {row.get('score', 0):.4f}")

    press_enter()


def menu_batch():
    """菜单 [2] 多疾病批量查询"""
    show_banner()
    print("\n  ▸ 多疾病批量查询")
    print("  " + "─" * 56)
    print("  输入多个疾病名称（每行一个），空行结束")
    print("  或输入预置编号（如 1,3,5）批量查询\n")
    _show_preset_diseases()
    print()

    raw = input("  请输入疾病名称或编号（每行一个，空行结束）:\n  (第一行) ").strip()
    if not raw:
        press_enter()
        return

    lines = [raw]
    while True:
        line = input("  ").strip()
        if not line:
            break
        lines.append(line)

    # 解析输入
    queries = []
    for item in lines:
        if item in PRESET_DISEASES:
            info = PRESET_DISEASES[item]
            queries.append((info["id"], info["name"]))
        else:
            queries.append((None, item))

    if not queries:
        press_enter()
        return

    print(f"\n  共 {len(queries)} 个查询任务，开始执行...\n")

    output_dir = _output_dir()
    for i, (did, keyword) in enumerate(queries, start=1):
        # 如果无 ID，先搜索
        if did is None:
            hits = search_disease(keyword)
            if hits:
                did = hits[0]["id"]
                logger.info(f"[{i}/{len(queries)}] 「{keyword}」→ {hits[0]['name']} ({did})")
            else:
                logger.warning(f"[{i}/{len(queries)}] 「{keyword}」→ 未找到")
                continue

        logger.info(f"[{i}/{len(queries)}] 正在获取靶点...")
        all_rows, dn = fetch_all_targets(did)
        if not all_rows:
            logger.warning(f"[{i}/{len(queries)}] 无靶点数据")
            continue

        all_rows.sort(key=lambda r: r.get("score") or 0, reverse=True)
        filepath = save_results(all_rows, dn or keyword, keyword)
        logger.info(f"[{i}/{len(queries)}] 完成: {len(all_rows)} 靶点 → {os.path.basename(filepath)}")

        if i < len(queries):
            time.sleep(RATE_LIMIT_DELAY)

    print(f"\n  ✓ 批量查询完成，文件保存在: {output_dir}")
    press_enter()


# ===================================================================
#  主入口
# ===================================================================

def main():
    while True:
        show_banner()
        print()
        print("  ┌──────────────────────────────────────────────────────┐")
        print("  │  [1] 搜索疾病并检索靶点    输入名称 → 选择 → 获取  │")
        print("  │  [2] 多疾病批量查询        一次检索多个疾病         │")
        print("  │  [0] 退出                                            │")
        print("  └──────────────────────────────────────────────────────┘")
        print()
        print("  提示: 还支持直接输入快捷编号（如 1=Alzheimer, 2=Diabetes）")
        print("        或直接粘贴疾病 ID（MONDO_xxx / EFO_xxx）跳过搜索")

        choice = input("\n  请选择 [0-2]: ").strip()

        if choice == "1":
            menu_search()
        elif choice == "2":
            menu_batch()
        elif choice == "0":
            show_banner()
            print("\n  感谢使用 Open Targets Platform 疾病靶点检索工具，再见！\n")
            break
        else:
            print(f"\n  ⚠ 无效选项「{choice}」，请重新选择")
            time.sleep(1)


if __name__ == "__main__":
    main()
